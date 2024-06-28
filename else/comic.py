import struct
import openpyxl
import os

f = open("resCmd.txt")
line = f.readlines()
cmdList = []

for l in line:
    res = l.split(", ")[0]
    cmdList.append(res)
f.close()

wb = openpyxl.load_workbook("dend.xlsx")

for file in os.listdir():
    if file.find("COMIC") != -1:
        index = 16
        f = open(file, "rb")
        wb.create_sheet(title=file[:-4])
        ws = wb[file[:-4]]
        line = f.read()
        index += 1

        #ReadComicImg
        imgCnt = line[index]
        index += 1
        for i in range(imgCnt):
            b = line[index]
            index += 1
            text = line[index:index+b].decode()
            print(text)
            index += b
        print()

        #ReadComicSize

        b = line[index]
        index += 1
        for i in range(b):
            index += 1
            for j in range(4):
                text = line[index:index+4]
                if struct.unpack("<f", text)[0] != -1:
                    print("ReadComicSize Error!")
                index += 4
        print()
        #ReadSE

        secnt = line[index]
        index += 1
        for i in range(secnt):
            b = line[index]
            index += 1
            text = line[index:index+b].decode()
            print(text)
            index += b
            index += 1
        print()

        #ReadBGM
        bgmcnt = line[index]
        index += 1
        for i in range(bgmcnt):
            b = line[index]
            index += 1
            text = line[index:index+b].decode()
            print(text)
            index += b
            index += 1
            index += 4
            index += 4
        print()

        #ReadComicData
        index += 1
        num = struct.unpack("<H", line[index:index+2])[0]
        index += 2
        FVTList = []
        for i in range(num):
            num2 = struct.unpack("<H", line[index:index+2])[0]
            index += 2
            b = line[index]
            if b >= 16:
                print("scriptErr!!")
                b = 16
            array = []
            p_str = []
            index += 1
            for j in range(b):
                array.append(struct.unpack("<f", line[index:index+4])[0])
                index += 4
            if num2 == 50:
                print("KomaCnt", array[0]+1)
            elif num2 != 51:
                if num2 == 193:
                    p_str.append(array[0])
                elif num2 == 54:
                    print("KomaCnt", array[0]+1)
                elif num2 == 252 or num2 == 68:
                    FVTList.append(array[0])
                elif num2 == 565:
                    FVTList.append(array[0])
                    FVTList.append(array[1])
                    FVTList.append(array[2])

            ws.cell(row=i+1, column=1).value = cmdList[num2].replace("'", "")
            for arr in range(len(array)):
                ws.cell(row=i+1, column=arr+2).value = array[arr]

        f.close()

wb.save("dend.xlsx")
